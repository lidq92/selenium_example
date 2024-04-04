# -*- coding: utf-8 -*-
"""
Created on 2020/11/24

python3

@author: Dingquan Li
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import urllib
import re
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import workbook 
from openpyxl import load_workbook
import argparse

parser = argparse.ArgumentParser()
parser.add_argument('--file', '-f', type=str, default='test.xlsx')
parser.add_argument('--start', '-s', type=int, default=0)
args = parser.parse_args()


def read_excel(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    
    titles = []
    
    for row in ws.iter_rows(min_row=3):
        if row[1].value is not None:
            titles.append(row[1].value)
    
    return titles
# def read_excel(file_path):
#     wb = load_workbook(file_path)
#     ws = wb.active
    
#     titles = []
#     urls = []
    
#     for row in ws.iter_rows(min_row=3):
#         if row[2].value is not None:
#             titles.append(row[2].value)
#             urls.append(row[9].value)
    
#     return titles, urls

titles = read_excel(args.file)
wb = load_workbook(args.file)
ws = wb[wb.sheetnames[0]]
print('Total number of journals: {}'.format(len(titles)))
browser = webdriver.Firefox() # webdriver.Chrome()
browser.maximize_window()
browser.get('https://www.letpub.com.cn/index.php?page=login')
# browser.find_element(By.ID, 'email').send_keys('youremail')
# browser.find_element(By.ID, 'password').send_keys('yourpassword')
# browser.find_element(By.XPATH, '//img[@onclick="login();"]').click()

browser.execute_script("window.open();")
browser.switch_to.window(browser.window_handles[1])

for i, title in enumerate(titles):
    if i < args.start:
        continue
    browser.get('https://www.letpub.com.cn/index.php?page=journalapp')
    if title == "CSIAM Transactions on Applied Mathematics     CSIAM应用数学会刊(英文版)":
        title = "CSIAM Transactions on Applied Mathematics" # 此期刊的title特殊处理
    title = re.sub(r'\s*[\u4e00-\u9fff].*$', '', title) # 去掉英文期刊的中文备注
    title = title.rstrip() # 去除表格中JOURNAL OF MATHEMATICAL BIOLOGY 等英文期刊最后出现的空格

    if title == "Probability, Uncertainty and Quantitative Risk":
        title = "Probability Uncertainty and Quantitative Risk" # letpub显示的是这个错误的名字，需要跟这个对上
    if title == "Biostatistics & Epidemiology":
        title = "Biostatistics and Epidemiology" # 

    print(title)
    if title == '': continue # 中文期刊在letpub查询不到
    browser.implicitly_wait(10)
    try:
        browser.find_element(By.CSS_SELECTOR, "[class='layui-layer-ico layui-layer-close layui-layer-close2']").click()
        print('Close the first pop-up!')
        browser.implicitly_wait(10)
        browser.find_element(By.CSS_SELECTOR, "[class='layui-layer-ico layui-layer-close layui-layer-close1']").click()
        print('Close the second pop-up!')
    except:
        pass
        # print('No need to close window pop-up!')
    browser.find_element(By.ID, 'searchname').send_keys(title)
    browser.find_element(By.XPATH, '''//input[@onclick="window.location='./index.php?page=journalapp&view=search&searchname='+escape(document.searchkeysform.searchname.value)+'&searchissn='+document.searchkeysform.searchissn.value+'&searchfield='+document.searchkeysform.searchfield.value+'&searchimpactlow='+document.searchkeysform.searchimpactlow.value+'&searchimpacthigh='+document.searchkeysform.searchimpacthigh.value+'&searchimpacttrend='+document.searchkeysform.searchimpacttrend.value+'&searchscitype='+document.searchkeysform.searchscitype.value+'&searchcategory1='+document.searchkeysform.searchcategory1.value+'&searchcategory2='+document.searchkeysform.searchcategory2.value+'&searchjcrkind='+document.searchkeysform.searchjcrkind.value+'&searchopenaccess='+document.searchkeysform.searchopenaccess.value+'&searchsort='+document.searchkeysform.searchsort.value; return false;"]''').click()
    xpath = f'//a[normalize-space(translate(text(), "ABCDEFGHIJKLMNOPQRSTUVWXYZ", "abcdefghijklmnopqrstuvwxyz")) = "{title.lower()}"]' # 处理期刊大小写不一致的问题
    try:
        element = browser.find_element(By.XPATH, xpath)
        relative_href = element.get_attribute('href')
        browser.get(relative_href)
        browser.implicitly_wait(2)
        # print('Read table')

        table = browser.find_elements(By.CLASS_NAME, 'table_yjfx')[1]
        # print(table.get_attribute('innerHTML'))
        rows = table.find_elements(By.TAG_NAME, "tr")
        for row in rows:
            col = row.find_elements(By.TAG_NAME, "td")
            # print(col.get_attribute('innerHTML'))
            if len(col) >= 2:
                if col[0].text == '期刊ISSN':
                    ws.cell(i+3, 4).value = col[1].text
                if col[0].text == 'E-ISSN':
                    ws.cell(i+3, 5).value = col[1].text
                # if col[0].text in ['期刊ISSN','E-ISSN']:
                #     print(col[0].text + ': ' + col[1].text)
    except:
        print('{} not found in letpub!'.format(title))
    # if i % 5 == 4: # 及时保存
    #     wb.save(args.file[:-5]+'_new.xlsx')
    wb.save(args.file[:-5]+'_new.xlsx')
browser.quit()

## https://portal.issn.org/ 其实只是通过题目查找ISSN的话，可以用这个网址

## pure math
# 24    Peking Mathematical Journal https://link.springer.com/journal/42543 ISSN=2096-6075 E-ISSN=2524-7182
# 119	Communications in Mathematical Research https://www.global-sci.org/cmr ISSN=1674-5647
# 147	Journal of Mathematical Research with Applications http://jmre.ijournals.cn/en/ch/first_menu.aspx?parent_id=20120223164315001 ISSN=2095-2651
# 170	PUBLICATIONES MATHEMATICAE-DEBRECEN https://publi.math.unideb.hu/ ISSN=0033-3883, E-ISSN 2064-2849
#
## applied math
# 30    ESAIM: Mathematical Modelling and Numerical Analysis (ESAIM: M2AN) https://www.esaim-m2an.org/ ISSN=2822-7840 E-ISSN=2804-7214
# 62    Annals of Applied Mathematics https://www.global-sci.com/aam.html ISSN=2096-0174
# 87    Journal of Nonlinear Modeling and Analysis https://www.global-sci.org/jnma.html ISSN=2562-2854
# 97	高等学校计算数学学报（中文版）http://maths.nju.edu.cn/~hjw/nmcn/c_index.php ISSN=1004-8979
# 98	高校应用数学学报A辑 http://www.amjcu.zju.edu.cn/ ISSN=1000-4424 https://portal.issn.org/resource/ISSN/1000-4424 
# 99	计算数学（中文版） https://computmath.cjoe.ac.cn/jssx/CN/home ISSN=0254-7791
# 100	应用数学 https://yisu.cbpt.cnki.net/WKB/WebPublication/index.aspx?mid=YISU ISSN=1001-9847
# 101	应用数学学报（中文版）https://applmath.cjoe.ac.cn/jweb_yysxxb/CN/home ISSN=0254-3079
# 152	应用泛函分析学报 http://yyfhfxxb.llyj.net/ ISSN=1009-1327
# 153	应用数学和力学（中文版）http://applmathmech.cqjtu.edu.cn/ ISSN=1000-0887

# 103	Advances in Applied Clifford Algebras https://link.springer.com/journal/6 ISSN=0188-7009 E-ISSN=1661-4909
# 106	ANNALS OF MATHEMATICS AND ARTIFICIAL INTELLIGENCE https://link.springer.com/journal/10472 
# 110	Applied and Computational Mathematics http://acmij.az/view.php?lang=en&menu=0
# 126	INFORMATICA
# 	
## prob-stat
# 
#
## interdisciplinary
#
#