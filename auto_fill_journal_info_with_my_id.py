# -*- coding: utf-8 -*-
"""
Created on 2020/11/24

python3

@author: Dingquan Li
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import urllib
import re
import xlwt
import xlrd
from xlutils.copy import copy
from openpyxl import workbook 
from openpyxl import load_workbook
import argparse

parser = argparse.ArgumentParser()
parser.add_argument('--file', '-f', type=str, default='test.xlsx')
parser.add_argument('--start', '-s', type=int, default=0)
args = parser.parse_args()


def read_excel(file):
    """
    https://zhuanlan.zhihu.com/p/38492442
    """
    rb = xlrd.open_workbook(filename=file) #, formatting_info=True

    # print(rb.sheet_names())

    sheet1 = rb.sheet_by_index(0)

    # print(sheet1)

    # print(sheet1.name,sheet1.nrows,sheet1.ncols)

    # rows = sheet1.row_values(2)

    # cols = sheet1.col_values(9)

    # print(rows)

    # print(cols)

    # print(sheet1.cell(1,9).value)

    # print(sheet1.cell_value(1,0))

    # print(sheet1.row(1)[0].value)

    return sheet1.col_values(2)[2:], sheet1.col_values(9)[2:]


titles, urls = read_excel(args.file)
wb = load_workbook(args.file)
ws = wb[wb.sheetnames[0]]
print(len(urls))
browser = webdriver.Firefox() # webdriver.Chrome()
browser.maximize_window()
browser.get('https://www.letpub.com.cn/index.php?page=login')
browser.find_element_by_id('email').send_keys('youremail')
browser.find_element_by_id('password').send_keys('yourpassward')
browser.find_element(By.XPATH, "//img[@onclick=' login()']").click()

browser.execute_script("window.open();")
browser.switch_to.window(browser.window_handles[1])

for i, url in enumerate(urls):
    if i < args.start:
        continue
    browser.get(url)
    print(url)
    browser.implicitly_wait(2)
    if 'letpub' in url:
        # browser.implicitly_wait(10)
        try:
            browser.find_element_by_css_selector("[class='layui-layer-ico layui-layer-close layui-layer-close2']").click()
            print('Close the first pop-up!')
        except:
            print('No need to close window pop-up!')

        # print('Read table')

        table = browser.find_elements(By.CLASS_NAME, 'table_yjfx')[1]
        # print(table.get_attribute('innerHTML'))
        rows = table.find_elements(By.TAG_NAME, "tr")
        for row in rows:
            col = row.find_elements(By.TAG_NAME, "td")
            # print(col.get_attribute('innerHTML'))
            if len(col) >= 2:
                if col[0].text == '是否OA开放访问':
                    ws.cell(i+3, 9).value = col[1].text
                if col[0].text == '出版商':
                    ws.cell(i+3, 5).value = col[1].text
                if col[0].text == '出版国家或地区':
                    ws.cell(i+3, 4).value = col[1].text
                if col[0].text == '出版周期':
                    ws.cell(i+3, 7).value = col[1].text
                if col[0].text == '出版年份':
                    if col[1].text == '0' or col[1].text == 0:
                        print("No info. about first published year")
                    else:
                        ws.cell(i+3, 6).value = col[1].text
                if col[0].text == '年文章数':
                    nwzs = col[1].text[:].replace('查看年文章数趋势图', '')
                    ws.cell(i+3, 8).value = nwzs.replace('点击', '')
                # if col[0].text in ['是否OA开放访问','出版商','出版国家或地区','出版周期','出版年份','年文章数']:
                #     print(col[0].text + ': ' + col[1].text)
    elif 'cnki' in url: 
        # browser.find_element_by_id('J_sumBtn-stretch').click()
        browser.find_element_by_xpath("//*[contains(text(), '更多介绍')]").click()
        zbdw = browser.find_elements_by_xpath("//p[contains(text(), '主办单位')]")
        ws.cell(i+3, 5).value = zbdw[0].text[5:]
        cbzq = browser.find_elements_by_xpath("//p[contains(text(), '出版周期')]")
        ws.cell(i+3, 7).value = cbzq[0].text[5:]
        cbd = browser.find_elements_by_xpath("//p[contains(text(), '出版地')]")
        ws.cell(i+3, 4).value = cbd[0].text[4:]
        cksj = browser.find_elements_by_xpath("//p[contains(text(), '创刊时间')]")
        ws.cell(i+3, 6).value = cksj[0].text[5:]
    else:
        print('Other url')
    if i % 10 == 9:
        wb.save(args.file[:-5]+'_new.xlsx')

wb.save(args.file[:-5]+'_new.xlsx')
browser.quit()
